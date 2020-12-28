### input的两组数据分为两个sheet读入
#引入对excel操作的包
import xlrd
import xlwt
import openpyxl
#读取表格
readbook = xlrd.open_workbook("input.xlsx")
#输出excel表格中的sheet名字
print (readbook.sheet_names())
time = [] #死亡天数
status = [] #1表示死亡个数
group = [] #处理组和对照组用输入1，2，3...表示
#打开名为'Sheet1'的sheet
sheet1 = readbook.sheet_by_name('Sheet1')
nrows = sheet1.nrows#行
ncols = sheet1.ncols#列
#输出行列值
print(nrows,ncols)
for i in range(nrows):
    a = (sheet1.cell(i,0).value)#第i行，第1列数值
    b = (sheet1.cell(i,1).value)#第i行，第2列数值
    if b == 0:#判断b是否为0
        continue;
    else:
        for x in range (int(b)):
            time.append(a)
            status.append(1)
            group.append(1)
sheet2 = readbook.sheet_by_name('Sheet2')
nrows = sheet2.nrows#行
ncols = sheet2.ncols#列
#输出行列值
print(nrows,ncols)
for i in range(nrows):
    a = (sheet2.cell(i,0).value)#第i行，第1列数值
    b = (sheet2.cell(i,1).value)#第i行，第2列数值
    if b == 0:#判断b是否为0
        continue;
    else:
        for x in range (int(b)):
            time.append(a)
            status.append(1)
            group.append(2)
total_list = [time,status,group]
mywb = openpyxl.Workbook()
sheet = mywb.active;  # 获取初始的sheet
sheet.title = 'summary'
for r in range(len(total_list)):
    for c in range(len(total_list[0])):
        sheet.cell(c + 1, r + 1).value = total_list[r][c] # excel中的行和列是从1开始计数的，所以需要+1
mywb.save("output.xlsx")
print("成功写入文件")
